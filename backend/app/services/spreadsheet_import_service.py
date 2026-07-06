from datetime import date, datetime
from decimal import Decimal
from typing import Any, BinaryIO

from openpyxl import load_workbook

from app.models.finance import FieldMapping, ImportPreview, MonthlyFinanceRecord


FIELD_ALIASES: dict[str, list[str]] = {
    "period": ["期间", "月份", "年月", "会计期间", "period"],
    "revenue": ["营业收入", "收入", "主营业务收入", "revenue"],
    "cost": ["营业成本", "成本", "主营业务成本", "cost"],
    "sales_expense": ["销售费用", "销售费", "sales_expense"],
    "admin_expense": ["管理费用", "管理费", "admin_expense"],
    "rd_expense": ["研发费用", "研发费", "rd_expense"],
    "finance_expense": ["财务费用", "财务费", "finance_expense"],
    "total_profit": ["利润总额", "total_profit"],
    "net_profit": ["净利润", "net_profit"],
    "cash": ["货币资金", "现金", "银行存款", "cash"],
    "accounts_receivable": ["应收账款", "应收款", "accounts_receivable"],
    "inventory": ["存货", "库存", "inventory"],
    "fixed_assets": ["固定资产", "fixed_assets"],
    "total_assets": ["资产总额", "总资产", "total_assets"],
    "short_term_loans": ["短期借款", "short_term_loans"],
    "accounts_payable": ["应付账款", "应付款", "accounts_payable"],
    "total_liabilities": ["负债总额", "总负债", "total_liabilities"],
    "owner_equity": ["所有者权益", "股东权益", "净资产", "owner_equity"],
    "operating_cash_flow_net": ["经营现金流净额", "经营活动现金流量净额", "经营现金流", "operating_cash_flow_net"],
    "operating_cash_inflow": ["经营现金流入", "经营活动现金流入", "operating_cash_inflow"],
    "operating_cash_outflow": ["经营现金流出", "经营活动现金流出", "operating_cash_outflow"],
    "investing_cash_flow_net": ["投资现金流净额", "投资活动现金流量净额", "investing_cash_flow_net"],
    "financing_cash_flow_net": ["筹资现金流净额", "筹资活动现金流量净额", "financing_cash_flow_net"],
    "customer_collection": ["客户回款", "回款金额", "customer_collection"],
    "sales_orders": ["销售订单", "销售订单金额", "sales_orders"],
    "purchase_amount": ["采购金额", "采购额", "purchase_amount"],
    "inventory_turnover_days": ["库存周转天数", "存货周转天数", "inventory_turnover_days"],
    "tax_burden_rate": ["税负率", "综合税负率", "tax_burden_rate"],
}

REQUIRED_FIELDS = ["period", "revenue", "cost", "net_profit", "operating_cash_flow_net"]

FIELD_LABELS: dict[str, str] = {
    "period": "期间",
    "revenue": "营业收入",
    "cost": "营业成本",
    "sales_expense": "销售费用",
    "admin_expense": "管理费用",
    "rd_expense": "研发费用",
    "finance_expense": "财务费用",
    "total_profit": "利润总额",
    "net_profit": "净利润",
    "cash": "货币资金",
    "accounts_receivable": "应收账款",
    "inventory": "存货",
    "fixed_assets": "固定资产",
    "total_assets": "资产总额",
    "short_term_loans": "短期借款",
    "accounts_payable": "应付账款",
    "total_liabilities": "负债总额",
    "owner_equity": "所有者权益",
    "operating_cash_inflow": "经营现金流入",
    "operating_cash_outflow": "经营现金流出",
    "operating_cash_flow_net": "经营现金流净额",
    "investing_cash_flow_net": "投资现金流净额",
    "financing_cash_flow_net": "筹资现金流净额",
    "customer_collection": "客户回款",
    "sales_orders": "销售订单",
    "purchase_amount": "采购金额",
    "inventory_turnover_days": "库存周转天数",
    "tax_burden_rate": "税负率",
}


def parse_finance_workbook(stream: BinaryIO) -> ImportPreview:
    workbook = load_workbook(stream, data_only=True, read_only=True)
    warnings: list[str] = []

    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        header_index, field_map = _find_header(rows)
        if header_index is None:
            continue

        records = _rows_to_records(rows[header_index + 1 :], field_map, warnings)
        if records:
            header_row = rows[header_index]
            matched_fields = [
                str(header_row[column_index]).strip()
                for column_index, field in field_map.items()
                if field in MonthlyFinanceRecord.model_fields
            ]
            return ImportPreview(
                sheet_name=sheet.title,
                records=records,
                matched_fields=matched_fields,
                field_mappings=_build_field_mappings(header_row, field_map),
                warnings=warnings,
            )

    raise ValueError("未能识别可用财务数据表，请确认表头包含期间、营业收入、营业成本、净利润和经营现金流净额。")


def _find_header(rows: list[tuple[Any, ...]]) -> tuple[int | None, dict[int, str]]:
    for row_index, row in enumerate(rows[:12]):
        field_map: dict[int, str] = {}
        for column_index, value in enumerate(row):
            normalized = _normalize_header(value)
            if not normalized:
                continue
            matched = _match_field(normalized)
            if matched:
                field_map[column_index] = matched

        if all(required in field_map.values() for required in REQUIRED_FIELDS):
            return row_index, field_map

    return None, {}


def _build_field_mappings(header_row: tuple[Any, ...], field_map: dict[int, str]) -> list[FieldMapping]:
    headers_by_field: dict[str, str] = {}
    for column_index, field in field_map.items():
        if column_index >= len(header_row):
            continue
        value = header_row[column_index]
        if value is not None:
            headers_by_field[field] = str(value).strip()

    mappings: list[FieldMapping] = []
    for field, label in FIELD_LABELS.items():
        source_header = headers_by_field.get(field)
        required = field in REQUIRED_FIELDS
        matched = source_header is not None
        if matched:
            status = "matched"
        elif required:
            status = "missing_required"
        else:
            status = "missing_optional"
        mappings.append(
            FieldMapping(
                field=field,
                label=label,
                source_header=source_header,
                required=required,
                matched=matched,
                status=status,
            )
        )
    return mappings


def _rows_to_records(
    rows: list[tuple[Any, ...]], field_map: dict[int, str], warnings: list[str]
) -> list[MonthlyFinanceRecord]:
    records: list[MonthlyFinanceRecord] = []
    for row_number, row in enumerate(rows, start=2):
        values: dict[str, Any] = {}
        for column_index, field_name in field_map.items():
            if column_index >= len(row):
                continue
            values[field_name] = row[column_index]

        if not any(value not in (None, "") for value in values.values()):
            continue

        try:
            record = _build_record(values)
        except ValueError as error:
            warnings.append(f"第 {row_number} 行跳过：{error}")
            continue

        records.append(record)

    return sorted(records, key=lambda item: item.period)


def _build_record(values: dict[str, Any]) -> MonthlyFinanceRecord:
    period = _parse_period(values.get("period"))
    if not period:
        raise ValueError("期间为空或格式无法识别")

    revenue = _number(values.get("revenue"))
    cost = _number(values.get("cost"))
    net_profit = _number(values.get("net_profit"))
    operating_cash_flow_net = _number(values.get("operating_cash_flow_net"))
    investing_cash_flow_net = _number(values.get("investing_cash_flow_net"))
    financing_cash_flow_net = _number(values.get("financing_cash_flow_net"))

    total_assets = _number(values.get("total_assets"))
    total_liabilities = _number(values.get("total_liabilities"))
    owner_equity = _number(values.get("owner_equity"))

    if total_assets == 0:
        total_assets = (
            _number(values.get("cash"))
            + _number(values.get("accounts_receivable"))
            + _number(values.get("inventory"))
            + _number(values.get("fixed_assets"))
        )
    if total_liabilities == 0:
        total_liabilities = _number(values.get("short_term_loans")) + _number(values.get("accounts_payable"))
    if owner_equity == 0:
        owner_equity = total_assets - total_liabilities

    return MonthlyFinanceRecord(
        period=period,
        revenue=revenue,
        cost=cost,
        sales_expense=_number(values.get("sales_expense")),
        admin_expense=_number(values.get("admin_expense")),
        rd_expense=_number(values.get("rd_expense")),
        finance_expense=_number(values.get("finance_expense")),
        total_profit=_number(values.get("total_profit"), default=net_profit),
        net_profit=net_profit,
        cash=_number(values.get("cash")),
        accounts_receivable=_number(values.get("accounts_receivable")),
        inventory=_number(values.get("inventory")),
        fixed_assets=_number(values.get("fixed_assets")),
        total_assets=total_assets,
        short_term_loans=_number(values.get("short_term_loans")),
        accounts_payable=_number(values.get("accounts_payable")),
        total_liabilities=total_liabilities,
        owner_equity=owner_equity,
        operating_cash_inflow=_number(
            values.get("operating_cash_inflow"),
            default=max(operating_cash_flow_net, 0),
        ),
        operating_cash_outflow=_number(
            values.get("operating_cash_outflow"),
            default=max(-operating_cash_flow_net, 0),
        ),
        operating_cash_flow_net=operating_cash_flow_net,
        investing_cash_flow_net=investing_cash_flow_net,
        financing_cash_flow_net=financing_cash_flow_net,
        customer_collection=_number(values.get("customer_collection"), default=max(operating_cash_flow_net, 0)),
        sales_orders=_number(values.get("sales_orders"), default=revenue),
        purchase_amount=_number(values.get("purchase_amount"), default=cost),
        inventory_turnover_days=_number(values.get("inventory_turnover_days")),
        tax_burden_rate=_rate(values.get("tax_burden_rate")),
    )


def _normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip().replace(" ", "").replace("\n", "")


def _match_field(header: str) -> str | None:
    lowered = header.lower()
    for field, aliases in FIELD_ALIASES.items():
        if lowered in [alias.lower().replace(" ", "") for alias in aliases]:
            return field
    return None


def _parse_period(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime | date):
        return value.strftime("%Y-%m")
    text = str(value).strip()
    if not text:
        return ""
    text = text.replace("年", "-").replace("月", "").replace("/", "-").replace(".", "-")
    parts = text.split("-")
    if len(parts) >= 2 and parts[0].isdigit() and parts[1].isdigit():
        return f"{int(parts[0]):04d}-{int(parts[1]):02d}"
    return text


def _number(value: Any, default: float = 0) -> float:
    if value in (None, ""):
        return float(default)
    if isinstance(value, int | float | Decimal):
        return float(value)
    text = str(value).strip().replace(",", "").replace("，", "")
    if text.endswith("%"):
        return float(text[:-1]) / 100
    return float(text)


def _rate(value: Any) -> float:
    rate = _number(value)
    return rate / 100 if rate > 1 else rate
