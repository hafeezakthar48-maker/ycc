from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


EXCEL_MIME_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
TEMPLATE_FILENAME = "finance-analysis-template.xlsx"

TEMPLATE_FIELDS: list[tuple[str, str]] = [
    ("期间", "必填，格式为 YYYY-MM，例如 2026-06。"),
    ("营业收入", "必填，单位：万元。"),
    ("营业成本", "必填，单位：万元。"),
    ("销售费用", "单位：万元。"),
    ("管理费用", "单位：万元。"),
    ("研发费用", "单位：万元。"),
    ("财务费用", "单位：万元。"),
    ("净利润", "必填，单位：万元。"),
    ("货币资金", "单位：万元。"),
    ("应收账款", "单位：万元。"),
    ("存货", "单位：万元。"),
    ("固定资产", "单位：万元。"),
    ("资产总额", "单位：万元；为空时系统会尝试按资产项目合计。"),
    ("短期借款", "单位：万元。"),
    ("应付账款", "单位：万元。"),
    ("负债总额", "单位：万元；为空时系统会尝试按负债项目合计。"),
    ("所有者权益", "单位：万元；为空时系统会尝试按资产总额减负债总额。"),
    ("经营现金流净额", "必填，单位：万元。"),
    ("投资现金流净额", "单位：万元，流出可填负数。"),
    ("筹资现金流净额", "单位：万元，流出可填负数。"),
    ("库存周转天数", "单位：天。"),
    ("税负率", "可填 0.038 或 3.8%，系统会统一识别。"),
]

EXAMPLE_ROW = [
    "2026-06",
    1286,
    880,
    78,
    62,
    28,
    15,
    146,
    402,
    428,
    482,
    828,
    2216,
    242,
    276,
    812,
    1404,
    92,
    -42,
    12,
    74,
    0.038,
]


def build_finance_template_workbook() -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "财务数据模板"
    sheet.freeze_panes = "A2"
    sheet.append([field for field, _description in TEMPLATE_FIELDS])
    sheet.append(EXAMPLE_ROW)

    _style_data_sheet(sheet)
    _build_instruction_sheet(workbook)

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _style_data_sheet(sheet) -> None:
    header_fill = PatternFill("solid", fgColor="1F6F8B")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for index, (field, _description) in enumerate(TEMPLATE_FIELDS, start=1):
        column_letter = get_column_letter(index)
        sheet.column_dimensions[column_letter].width = max(len(field) * 2.2, 14)

    sheet["V2"].number_format = "0.00%"


def _build_instruction_sheet(workbook: Workbook) -> None:
    sheet = workbook.create_sheet("填报说明")
    sheet["A1"] = "填报说明"
    sheet["A1"].font = Font(bold=True, size=14)
    sheet.append(["字段", "说明"])

    for field, description in TEMPLATE_FIELDS:
        sheet.append([field, description])

    for cell in sheet[2]:
        cell.fill = PatternFill("solid", fgColor="E8F3FF")
        cell.font = Font(bold=True)

    sheet.column_dimensions["A"].width = 22
    sheet.column_dimensions["B"].width = 64
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
